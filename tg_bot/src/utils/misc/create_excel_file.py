import io
from itertools import zip_longest

import openpyxl
from openpyxl.styles import Font


def create_bytes_excel_file(data: list, headers: list) -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    bold_font = Font(bold=True)

    for i, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=i)
        cell.value = header
        cell.font = bold_font

    max_columns = max(len(headers), max(len(row_data) for row_data in data))

    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_data in zip_longest(range(1, max_columns + 1), row_data.values(), fillvalue="N/A"):
            sheet.cell(row=row_num, column=col_num).value = cell_data

    with io.BytesIO() as file_stream:
        workbook.save(file_stream)
        return file_stream.getvalue()
